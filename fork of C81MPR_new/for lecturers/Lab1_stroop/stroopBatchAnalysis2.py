from psychopy import gui, core, misc
from numpy import * #lots of numeric things, like mean and std
import os

congRTcells = ['R2','R3', 'R7','R8', 'R12','R13']#make sure these are right!
incongRTcells = ['R4','R5','R6', 'R9','R10','R11']
groupName = 'group4reversestroop'
outputFilename = groupName+'.xlsx'
anonymous = True #if false then filenames will be added to sheet

#create excel file for output
from openpyxl.writer.excel import ExcelWriter
from openpyxl.workbook import Workbook
xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]#use the first worksheet (index 0)
xlSheet.title = groupName
xlSheet.cell('A1').value = 'congruent'
xlSheet.cell('B1').value = 'incongruent'

#use a file open dialog to choose the files to include
files = gui.fileOpenDlg(tryFilePath=".", allowed='*.psydat')
if not files:#user pressed cancel
    core.quit()
    
#do the actual analysis, file-by-file
row=2
for fileName in files:
    try:
        subj = misc.fromFile(fileName)
        
        #get all the RTs for each condition for this subject
        congRTs = []#start with an empty list
        incongRTs = []
        nTypes = len(subj.trialList)
        for condN, cond in enumerate(subj.trialList):
            rt = subj.data['resp.rt'][condN]
            if cond['congruent']=='cong':
                congRTs.append(rt)
            else:
                incongRTs.append(rt)
        #get mean RTs for each condition for this subject
        meanCongRT = mean(congRTs)
        meanIncongRT = mean(incongRTs)
        xlSheet.cell('A'+str(row)).value = meanCongRT
        xlSheet.cell('B'+str(row)).value = meanIncongRT
        if not anonymous:
            name = subj.extraInfo['participant']
            xlSheet.cell('C'+str(row)).value = name
        print meanCongRT, meanIncongRT
        row=row+1#go t onext row
    except:
        print 'failed to analyse', fileName

xlsxWriter.save(filename = outputFilename)