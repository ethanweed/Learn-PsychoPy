from psychopy import gui, core, misc
from numpy import * #lots of numeric things, like mean and std
import os

"""this is like the batch analysis for stroop, except that we now need to dig the age and gender out of the info
as well for the sake of the report.
"""
#libs for handling excel files:
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

groupName = 'group4x'
outputFilename = groupName+'.xlsx'
genderCell='B21'#make sure these are right!
ageCell = 'B22'
congRTcells = ['R2','R4', 'R6','R8','R10','R12','R14','R16']
incongRTcells = ['R3','R5','R7','R9','R11','R13','R15','R17']
anonymous = True #if false then filenames will be added to sheet

#use a file open dialog to choose the files to include
files = gui.fileOpenDlg(tryFilePath=".", allowed='*.psydat')
if not files:#user pressed cancel
    core.quit()

xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]#use the first worksheet (index 0)
xlSheet.title = groupName

xlSheet.cell('A1').value = 'gender'
xlSheet.cell('B1').value = 'age'
xlSheet.cell('C1').value = 'congruent'
xlSheet.cell('D1').value = 'incongruent'
row = 2#starting row for data
#do the actual analysis, file-by-file
for fileName in files:
    name = os.path.split(fileName)[1]#just the filename not whole path

#    try: #just ignore the whole file if there are errors

    subj = misc.fromFile(fileName)

    #get all the RTs for each condition for this subject
    congRTs = []#start with an empty list
    incongRTs = []
    nTypes = len(subj.trialList)
    for condN, cond in enumerate(subj.trialList):
        rt = subj.data['resp.rt'][condN]
        if cond['congruence']=='consistent':
            congRTs.append(rt)
        else:
            incongRTs.append(rt)
    print subj.extraInfo
    xlSheet.cell('A'+str(row)).value = subj.extraInfo['gender (m/f)']
    xlSheet.cell('B'+str(row)).value = subj.extraInfo['age']
    xlSheet.cell('C'+str(row)).value = mean(congRTs)
    xlSheet.cell('D'+str(row)).value = mean(incongRTs)
    xlSheet.cell('E'+str(row)).value = subj.extraInfo['participant']
#    print gender, age, meanIncongRT, meanCongRT
    row=row+1#go t onext row
#    except:
#        print "Problem analysing file ", name

xlsxWriter.save(filename = outputFilename)