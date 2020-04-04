
from psychopy import gui, core, misc
from numpy import * #lots of numeric things, like mean and std
import os

"""
Batch analysis script for Sternberg experiment. This script is a lot more
advanced than the one for analysing the Stroop task. Rather than
hard-coding the location of each RT location in the spreadsheet you
only need to  check the columns for the input files still match. The script
will use these to determine and combine the conditions automatically.

As a result, this script is immune to reordering of the rows (but not the columns)
whereas the Stroop batch analysis needs the cells to be in exactly the 'right' places
"""

groupName = 'group4'
outputFilename = groupName+'.xlsx'
anonymous = True #if false then filenames will be added to sheet

#columns of input files
ageCol = 'F' #the column storing "was the target present (y/n)"
genderCol = 'E' #the column storing "Set size (integer)"
digitRatioCol= 'H' #the column storing "did they get it correct (1/0)"
groupCol  = 'L' #the column of mean RTs
bartCol = 'M'
earnings = 'N'

#columns of output file
outCols={'age':'B',
        'gender':'C',
        'digitRatio':'D',
        'group':'E',
        'bart':'F',
        }

#libs for handling excel files:
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

#use a file open dialog to choose the files to include
files = gui.fileOpenDlg(tryFilePath=".", allowed='*.psydat')
if not files:#user pressed cancel
    core.quit()

xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]#use the first worksheet (index 0)
xlSheet.title = groupName
#make a header row
for condition in outCols.keys():
    xlSheet.cell(outCols[condition]+'1').value = condition

outRow = 2#starting row for data
#do the actual analysis, file-by-file
condNames = []
for fileName in files:
#    try:
        #create a python dict for each set of data
        pumps=[]
        earnings=[]
        subj = misc.fromFile(fileName)

        #process each row from this subject's data file
        inputRow = 2
        dataEnded=False
        for condN, cond in enumerate(subj.trialList):
            #about this condition
            if subj.data['popped'][condN]:
                continue
            earnings.append(subj.data['earnings'][condN])
            pumps.append(subj.data['nPumps'][condN])
            
        totEarnings=sum(earnings)
        meanPumps=mean(pumps)
        name = subj.extraInfo['participant']
        print 'analysed %s' %(name)
        
        xlSheet.cell(outCols['age']+str(outRow)).value = subj.extraInfo['age']
        if subj.extraInfo['gender (m/f)'] in ['m','M']:
            xlSheet.cell(outCols['gender']+str(outRow)).value = 0
        else:xlSheet.cell(outCols['gender']+str(outRow)).value = 1
        xlSheet.cell(outCols['digitRatio']+str(outRow)).value = subj.extraInfo['digit ratio']
        if subj.extraInfo['group (A/B)'] in ['a','A']:
            xlSheet.cell(outCols['group']+str(outRow)).value = 0
        else:xlSheet.cell(outCols['group']+str(outRow)).value = 1
        xlSheet.cell(outCols['bart']+str(outRow)).value = meanPumps
        if not anonymous:
            xlSheet.cell('A'+str(outRow)).value = name
        outRow=outRow+1#goto next row for next subject
#    except:
#        print 'problem with', fileName
xlsxWriter.save(filename = outputFilename)
