from psychopy import gui, core, misc
from numpy import * #lots of numeric things, like mean and std
import os

"""
Batch analysis script for Sternberg experiment. This script is a lot more
advanced than the one for analysing the Stroop task. Rather than
hard-coding the location of each RT location in the spreadsheet you 
only need to  check the columns for the input files still match. The script
will use these to determine and combine the conditions automatically.

As a result, this script is immune to reordering of the rows (but not the columns)
whereas the Stroop batch analysis needs the cells to be in exactly the 'right' places
"""

groupName = 'group4'
outputFilename = groupName+'.xlsx'
anonymous = True #if false then filenames will be added to sheet

#columns of input files
presentCol = 'F' #the column storing "was the target present (y/n)"
setSizeCol = 'E' #the column storing "Set size (integer)"
respAccCol= 'H' #the column storing "did they get it correct (1/0)"
respRTCol  = 'L' #the column of mean RTs

#columns of output file
outCols={'set1_y':'B',
        'set2_y':'C',
        'set3_y':'D',
        'set4_y':'E',
        'set5_y':'F',
        'set6_y':'G',
        'set1_n':'H',
        'set2_n':'I',
        'set3_n':'J',
        'set4_n':'K',
        'set5_n':'L',
        'set6_n':'M',
        }

#libs for handling excel files:
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

#use a file open dialog to choose the files to include
files = gui.fileOpenDlg(tryFilePath=".", allowed='*.psydat')
if not files:#user pressed cancel
    core.quit()
    
xlBook = Workbook()
xlsxWriter = ExcelWriter(workbook = xlBook)
xlSheet = xlBook.worksheets[0]#use the first worksheet (index 0)
xlSheet.title = groupName
#make a header row
for condition in outCols.keys():
    xlSheet.cell(outCols[condition]+'1').value = condition
    
outRow = 2#starting row for data
#do the actual analysis, file-by-file
condNames = []
for fileName in files:
#    try:
        #create a python dict for each set of data
        RTs={}
        for condition in outCols.keys(): 
            RTs[condition]=[]#a list for each condition that we'll append data to
            
        subj = misc.fromFile(fileName)
        
        #process each row from this subject's data file
        inputRow = 2
        dataEnded=False
        for condN, cond in enumerate(subj.trialList):
            #about this condition
            thisSetSize=cond['setSize']
            thisPresent= cond['present']
            thisCondition = "set%i_%s" %(thisSetSize, thisPresent) #e.g. set6_y
            #about this response
            thisRT= subj.data['resp.rt'][condN]
            thisAcc= subj.data['resp.corr'][condN]
            #add this RT only if correct ans given
            if thisAcc==1 and thisRT<2.0:
                RTs[thisCondition].append(thisRT)
        name = subj.extraInfo['participant']
        print 'analysed %s' %(name)
        
        #now calculate means for each condition and insert into sheet
        meanRTs={}
        for condition in RTs.keys():
            thisMean = mean(RTs[condition])
            col=outCols[condition]
            xlSheet.cell(col+str(outRow)).value=str(thisMean)
            
        if not anonymous:
            xlSheet.cell('A'+str(row)).value = str(name)
        outRow=outRow+1#goto next row for next subject
#    except:
#        print 'problem with', fileName
xlsxWriter.save(filename = outputFilename)